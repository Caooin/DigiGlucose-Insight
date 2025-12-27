import traceback
from fastapi import APIRouter, HTTPException, Depends, status, Request
from sqlalchemy.orm import Session
from datetime import timedelta, datetime
import json
import csv
import io
import re
import random
from typing import Optional, List

from . import schemas, mcp, orchestrator, database, models, auth, auth_schemas, email_service

router = APIRouter()

# 简易的内存字典：session_id -> ConversationBufferMemory
session_memories = {}

# 创建Orchestrator实例
orchestrator_agent = orchestrator.OrchestratorAgent()


def get_memory(session_id: str):
    if session_id not in session_memories:
        session_memories[session_id] = mcp.create_memory()
    return session_memories[session_id]


# ==================== 认证相关API ====================

def validate_email(email: str) -> bool:
    """验证邮箱格式"""
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return bool(re.match(pattern, email))


def generate_verification_code() -> str:
    """生成6位数字验证码"""
    return ''.join([str(random.randint(0, 9)) for _ in range(6)])


@router.post("/auth/send-verify-code")
async def send_verify_code(
    request: Request,
    email_data: auth_schemas.EmailVerifyRequest,
    db: Session = Depends(database.get_db)
):
    """获取邮箱验证码"""
    try:
        email = email_data.email.strip().lower()
        
        # 1. 邮箱格式校验
        if not validate_email(email):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="邮箱格式不正确"
            )
        
        # 2. 检查邮箱是否已注册
        existing_user = db.query(models.User).filter(
            models.User.email == email
        ).first()
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="该邮箱已注册，请更换邮箱或找回密码"
            )
        
        # 3. 防刷限制：同一邮箱1分钟内只能获取1次
        now = datetime.utcnow()
        recent_code = db.query(models.EmailVerifyCode).filter(
            models.EmailVerifyCode.email == email,
            models.EmailVerifyCode.create_time > now - timedelta(minutes=1)
        ).first()
        if recent_code:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail="请稍后再试，1分钟内只能获取1次验证码"
            )
        
        # 4. 防刷限制：同一邮箱24小时内最多获取5次
        codes_24h = db.query(models.EmailVerifyCode).filter(
            models.EmailVerifyCode.email == email,
            models.EmailVerifyCode.create_time > now - timedelta(hours=24)
        ).count()
        if codes_24h >= 5:
            raise HTTPException(
                status_code=status.HTTP_429_TOO_MANY_REQUESTS,
                detail="24小时内获取验证码次数过多，请稍后再试"
            )
        
        # 5. 删除该邮箱未过期但未使用的旧验证码（避免混淆）
        db.query(models.EmailVerifyCode).filter(
            models.EmailVerifyCode.email == email,
            models.EmailVerifyCode.is_used == False,
            models.EmailVerifyCode.expire_time > now
        ).delete()
        
        # 6. 生成验证码
        code = generate_verification_code()
        expire_time = now + timedelta(minutes=10)  # 10分钟有效期
        
        # 7. 获取客户端IP（防刷）
        client_ip = request.client.host if request.client else None
        
        # 8. 存储验证码
        verify_code = models.EmailVerifyCode(
            email=email,
            code=code,
            expire_time=expire_time,
            is_used=False,
            ip_address=client_ip,
            create_time=now
        )
        db.add(verify_code)
        db.commit()
        
        # 9. 发送验证码邮件
        email_sent = email_service.send_verification_code_email(email, code)
        if not email_sent:
            # 如果邮件发送失败，删除已保存的验证码记录
            db.delete(verify_code)
            db.commit()
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="验证码发送失败，请稍后重试"
            )
        
        return {
            "message": "验证码已发送至您的邮箱，请查收",
            "expire_minutes": 10
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"发送验证码错误: {str(e)}")
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"发送验证码失败: {str(e)}"
        )


@router.post("/auth/register", response_model=auth_schemas.Token)
async def register(
    user_data: auth_schemas.UserRegister,
    db: Session = Depends(database.get_db)
):
    """用户注册（带验证码校验）"""
    try:
        email = user_data.email.strip().lower()
        verify_code = user_data.verify_code.strip()
        
        # 1. 基础表单校验
        # 邮箱格式校验
        if not validate_email(email):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="邮箱格式不正确"
            )
        
        # 密码强度校验（至少8位，包含大小写字母和数字）
        if len(user_data.password) < 8:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="密码长度至少8位"
            )
        if not re.search(r'[a-z]', user_data.password) or not re.search(r'[A-Z]', user_data.password) or not re.search(r'\d', user_data.password):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="密码必须包含大小写字母和数字"
            )
        
        # 验证码格式校验（6位数字）
        if not re.match(r'^\d{6}$', verify_code):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="验证码格式不正确"
            )
        
        # 2. 邮箱唯一性校验（双重保障）
        existing_user = db.query(models.User).filter(
            models.User.username == user_data.username
        ).first()
        if existing_user:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="用户名已存在"
            )
        
        existing_email = db.query(models.User).filter(
            models.User.email == email
        ).first()
        if existing_email:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="该邮箱已注册，请更换邮箱或找回密码"
            )
        
        # 3. 验证码有效性校验
        now = datetime.utcnow()
        code_record = db.query(models.EmailVerifyCode).filter(
            models.EmailVerifyCode.email == email,
            models.EmailVerifyCode.code == verify_code,
            models.EmailVerifyCode.is_used == False,
            models.EmailVerifyCode.expire_time > now
        ).first()
        
        if not code_record:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="验证码无效、已过期或未获取，请重新获取"
            )
        
        # 4. 校验通过，创建新用户
        hashed_password = auth.get_password_hash(user_data.password)
        new_user = models.User(
            username=user_data.username,
            email=email,
            hashed_password=hashed_password,
            email_verified=True  # 注册时已验证邮箱
        )
        db.add(new_user)
        db.flush()  # 先flush获取user_id
        
        # 5. 标记验证码为已使用
        code_record.is_used = True
        db.commit()
        db.refresh(new_user)
        
        # 6. 生成token
        access_token = auth.create_access_token(data={"sub": str(new_user.id)})
        
        return {
            "access_token": access_token,
            "token_type": "bearer",
            "user_id": new_user.id,
            "username": new_user.username
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"注册错误: {str(e)}")
        traceback.print_exc()
        db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"注册失败: {str(e)}"
        )


@router.post("/auth/login", response_model=auth_schemas.Token)
async def login(
    user_data: auth_schemas.UserLogin,
    db: Session = Depends(database.get_db)
):
    """用户登录"""
    try:
        print(f"登录请求 - 用户名: {user_data.username}")
        
        # 检查数据库连接
        try:
            # 查找用户
            user = db.query(models.User).filter(
                models.User.username == user_data.username
            ).first()
            print(f"数据库查询完成，用户: {user.username if user else 'None'}")
        except Exception as db_error:
            print(f"数据库查询错误: {str(db_error)}")
            print(traceback.format_exc())
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="数据库连接错误，请稍后重试"
            )
        
        if not user:
            print(f"用户不存在: {user_data.username}")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码错误",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        # 检查hashed_password是否存在
        if not user.hashed_password:
            print(f"用户密码哈希为空: {user_data.username}")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码错误",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        # 验证密码
        try:
            password_valid = auth.verify_password(user_data.password, user.hashed_password)
            print(f"密码验证结果: {password_valid}")
        except Exception as pwd_error:
            print(f"密码验证过程出错: {str(pwd_error)}")
            print(traceback.format_exc())
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="密码验证失败，请稍后重试"
            )
        
        if not password_valid:
            print(f"密码验证失败: {user_data.username}")
            raise HTTPException(
                status_code=status.HTTP_401_UNAUTHORIZED,
                detail="用户名或密码错误",
                headers={"WWW-Authenticate": "Bearer"},
            )
        
        # 生成token（JWT的subject必须是字符串）
        try:
            access_token = auth.create_access_token(data={"sub": str(user.id)})
            print(f"Token生成成功")
        except Exception as token_error:
            print(f"Token生成错误: {str(token_error)}")
            print(traceback.format_exc())
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Token生成失败，请稍后重试"
            )
        
        print(f"登录成功 - 用户ID: {user.id}, 用户名: {user.username}")
        
        return {
            "access_token": access_token,
            "token_type": "bearer",
            "user_id": user.id,
            "username": user.username
        }
    except HTTPException:
        raise
    except Exception as e:
        error_msg = str(e)
        print(f"登录过程发生未知错误: {error_msg}")
        print(traceback.format_exc())
        # 返回更友好的错误信息，但不在detail中暴露敏感信息
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="登录服务暂时不可用，请稍后重试"
        )


@router.get("/auth/me", response_model=auth_schemas.UserInfo)
async def get_current_user_info(
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """获取当前用户信息"""
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "age": current_user.age,
        "gender": current_user.gender,
        "diagnosis_type": current_user.diagnosis_type,
        "fasting_target_min": current_user.fasting_target_min,
        "fasting_target_max": current_user.fasting_target_max,
        "post_meal_target_max": current_user.post_meal_target_max,
    }


@router.post("/chat", response_model=schemas.ChatResponse)
async def chat_with_assistant(
    request: schemas.ChatRequest,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    try:
        # 使用认证用户
        user = current_user
        
        # 使用Orchestrator处理消息
        result = orchestrator_agent.process_message(
            user_id=user.id,
            session_id=request.session_id,
            message=request.message
        )
        
        return schemas.ChatResponse(reply=result["reply"])
    except Exception as e:
        # 打印详细错误信息到控制台
        print(f"Error in chat endpoint: {str(e)}")
        print(traceback.format_exc())
        # 返回友好的错误信息
        raise HTTPException(
            status_code=500,
            detail=f"处理请求时出错: {str(e)}"
        )


@router.get("/user-id")
async def get_user_id(
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """获取当前用户ID"""
    return {"user_id": current_user.id}


@router.get("/healthz")
async def health_check():
    return {"status": "ok"}


@router.get("/test-llm")
async def test_llm():
    """测试云雾API连接"""
    try:
        from langchain_openai import ChatOpenAI
        import os
        from dotenv import load_dotenv
        
        load_dotenv()
        
        # 使用相同的配置创建LLM实例
        api_key = os.getenv("YUNWU_API_KEY", "sk-YyCYXQPvPUKViIaX8wrbtjVatojJh9L25Ov82Mh36QuS0e6V")
        base_url = os.getenv("YUNWU_BASE_URL", "https://api.yunwu.ai/v1")
        
        test_llm = ChatOpenAI(
            model="gpt-4",  # 使用 gpt-4，云雾API可能不支持 gpt-3.5-turbo
            temperature=0,
            api_key=api_key,
            base_url=base_url,
        )
        
        # 测试简单的LLM调用
        response = test_llm.invoke("你好，请回复'测试成功'")
        
        return {
            "status": "success",
            "api_key_preview": api_key[:20] + "..." if len(api_key) > 20 else api_key,
            "base_url": base_url,
            "model": test_llm.model_name,
            "response": response.content if hasattr(response, 'content') else str(response)
        }
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        print(f"Test LLM Error: {str(e)}")
        print(error_trace)
        return {
            "status": "error",
            "error": str(e),
            "traceback": error_trace
        }


@router.get("/users/glucose-readings")
async def get_glucose_readings(
    limit: int = 50,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取当前用户的血糖记录"""
    readings = db.query(models.GlucoseReading).filter(
        models.GlucoseReading.user_id == current_user.id
    ).order_by(models.GlucoseReading.timestamp.desc()).limit(limit).all()
    
    return {
        "readings": [
            {
                "id": r.id,
                "value": r.value,
                "unit": r.unit,
                "timestamp": r.timestamp.isoformat(),
                "context": r.context,
                "meal_type": r.meal_type,
                "risk_level": r.risk_level,
            }
            for r in readings
        ]
    }


@router.get("/users/weekly-report")
async def get_weekly_report(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取当前用户的周报"""
    from .weekly_report import generate_weekly_report
    
    result = generate_weekly_report(current_user.id)
    return result


@router.get("/users/profile")
async def get_user_profile(
    current_user: models.User = Depends(auth.get_current_active_user)
):
    """获取当前用户档案"""
    return {
        "id": current_user.id,
        "username": current_user.username,
        "email": current_user.email,
        "age": current_user.age,
        "gender": current_user.gender,
        "diagnosis_type": current_user.diagnosis_type,
        "fasting_target_min": current_user.fasting_target_min,
        "fasting_target_max": current_user.fasting_target_max,
        "post_meal_target_max": current_user.post_meal_target_max,
    }


@router.put("/users/profile")
async def update_user_profile(
    profile: dict,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """更新当前用户档案"""
    # 更新字段
    if "age" in profile:
        current_user.age = profile["age"]
    if "gender" in profile:
        current_user.gender = profile["gender"]
    if "diagnosis_type" in profile:
        current_user.diagnosis_type = profile["diagnosis_type"]
    if "fasting_target_min" in profile:
        current_user.fasting_target_min = profile["fasting_target_min"]
    if "fasting_target_max" in profile:
        current_user.fasting_target_max = profile["fasting_target_max"]
    if "post_meal_target_max" in profile:
        current_user.post_meal_target_max = profile["post_meal_target_max"]
    if "email" in profile:
        current_user.email = profile["email"]
    
    db.commit()
    db.refresh(current_user)
    return {"message": "更新成功"}


# ==================== 数据可视化API ====================

@router.get("/users/glucose-visualization")
async def get_glucose_visualization(
    days: int = 7,
    context: Optional[str] = None,  # fasting, post_meal, etc.
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取血糖可视化数据"""
    from datetime import datetime, timedelta
    
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    query = db.query(models.GlucoseReading).filter(
        models.GlucoseReading.user_id == current_user.id,
        models.GlucoseReading.timestamp >= start_date,
        models.GlucoseReading.timestamp <= end_date
    )
    
    if context:
        query = query.filter(models.GlucoseReading.context == context)
    
    readings = query.order_by(models.GlucoseReading.timestamp.asc()).all()
    
    # 格式化数据
    chart_data = {
        "dates": [],
        "values": [],
        "contexts": [],
        "meal_types": [],
        "risk_levels": []
    }
    
    for reading in readings:
        chart_data["dates"].append(reading.timestamp.isoformat())
        chart_data["values"].append(reading.value)
        chart_data["contexts"].append(reading.context or "")
        chart_data["meal_types"].append(reading.meal_type or "")
        chart_data["risk_levels"].append(reading.risk_level or "")
    
    # 计算统计数据
    if readings:
        values = [r.value for r in readings]
        stats = {
            "average": sum(values) / len(values),
            "max": max(values),
            "min": min(values),
            "count": len(readings)
        }
    else:
        stats = {"average": 0, "max": 0, "min": 0, "count": 0}
    
    return {
        "chart_data": chart_data,
        "stats": stats,
        "period": {"start": start_date.isoformat(), "end": end_date.isoformat(), "days": days}
    }


@router.get("/users/glucose-comparison")
async def get_glucose_comparison(
    days: int = 7,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取空腹/餐后血糖对比数据"""
    from datetime import datetime, timedelta
    
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    readings = db.query(models.GlucoseReading).filter(
        models.GlucoseReading.user_id == current_user.id,
        models.GlucoseReading.timestamp >= start_date,
        models.GlucoseReading.timestamp <= end_date
    ).order_by(models.GlucoseReading.timestamp.asc()).all()
    
    # 按类型分组
    fasting_readings = [r for r in readings if r.context == "fasting"]
    post_meal_readings = [r for r in readings if r.context == "post_meal"]
    
    # 计算每日平均值
    daily_data = {}
    for reading in readings:
        date_key = reading.timestamp.date().isoformat()
        if date_key not in daily_data:
            daily_data[date_key] = {"fasting": [], "post_meal": []}
        
        if reading.context == "fasting":
            daily_data[date_key]["fasting"].append(reading.value)
        elif reading.context == "post_meal":
            daily_data[date_key]["post_meal"].append(reading.value)
    
    # 计算每日平均
    comparison_data = {
        "dates": [],
        "fasting_avg": [],
        "post_meal_avg": []
    }
    
    for date_key in sorted(daily_data.keys()):
        comparison_data["dates"].append(date_key)
        fasting_vals = daily_data[date_key]["fasting"]
        post_meal_vals = daily_data[date_key]["post_meal"]
        comparison_data["fasting_avg"].append(
            sum(fasting_vals) / len(fasting_vals) if fasting_vals else None
        )
        comparison_data["post_meal_avg"].append(
            sum(post_meal_vals) / len(post_meal_vals) if post_meal_vals else None
        )
    
    # 总体统计
    stats = {
        "fasting": {
            "average": sum([r.value for r in fasting_readings]) / len(fasting_readings) if fasting_readings else 0,
            "count": len(fasting_readings)
        },
        "post_meal": {
            "average": sum([r.value for r in post_meal_readings]) / len(post_meal_readings) if post_meal_readings else 0,
            "count": len(post_meal_readings)
        }
    }
    
    return {
        "comparison_data": comparison_data,
        "stats": stats
    }


@router.get("/users/diet-visualization")
async def get_diet_visualization(
    days: int = 7,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取饮食可视化数据（高GI/低GI占比）"""
    from datetime import datetime, timedelta
    
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    meals = db.query(models.MealEntry).filter(
        models.MealEntry.user_id == current_user.id,
        models.MealEntry.timestamp >= start_date,
        models.MealEntry.timestamp <= end_date
    ).all()
    
    # 分类统计
    high_gi_count = 0
    low_gi_count = 0
    unknown_count = 0
    
    for meal in meals:
        if meal.estimated_gi:
            if meal.estimated_gi >= 70:
                high_gi_count += 1
            elif meal.estimated_gi < 70:
                low_gi_count += 1
        else:
            unknown_count += 1
    
    total = len(meals)
    
    return {
        "diet_data": {
            "high_gi": high_gi_count,
            "low_gi": low_gi_count,
            "unknown": unknown_count,
            "total": total
        },
        "percentages": {
            "high_gi": (high_gi_count / total * 100) if total > 0 else 0,
            "low_gi": (low_gi_count / total * 100) if total > 0 else 0,
            "unknown": (unknown_count / total * 100) if total > 0 else 0
        }
    }


# ==================== 趋势分析API ====================

@router.get("/users/glucose-trend")
async def get_glucose_trend(
    days: int = 7,
    context: Optional[str] = None,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取血糖趋势分析数据"""
    from datetime import datetime, timedelta
    import statistics
    
    end_date = datetime.utcnow()
    start_date = end_date - timedelta(days=days)
    
    query = db.query(models.GlucoseReading).filter(
        models.GlucoseReading.user_id == current_user.id,
        models.GlucoseReading.timestamp >= start_date,
        models.GlucoseReading.timestamp <= end_date
    )
    
    if context:
        query = query.filter(models.GlucoseReading.context == context)
    
    readings = query.order_by(models.GlucoseReading.timestamp.asc()).all()
    
    if not readings:
        return {
            "trend": "no_data",
            "trend_direction": "stable",
            "average_change": 0,
            "interpretation": "暂无数据",
            "chart_data": {"dates": [], "values": [], "trend_line": [], "anomalies": []},
            "stats": {}
        }
    
    # 计算趋势
    values = [r.value for r in readings]
    dates = [r.timestamp for r in readings]
    
    # 简单线性回归计算趋势
    n = len(values)
    x = list(range(n))
    x_mean = sum(x) / n
    y_mean = sum(values) / n
    
    numerator = sum((x[i] - x_mean) * (values[i] - y_mean) for i in range(n))
    denominator = sum((x[i] - x_mean) ** 2 for i in range(n))
    
    slope = numerator / denominator if denominator != 0 else 0
    intercept = y_mean - slope * x_mean
    
    # 趋势线数据
    trend_line = [slope * i + intercept for i in range(n)]
    
    # 标记异常值（超出正常范围的值）
    anomalies = []
    target_max = current_user.post_meal_target_max or 10.0
    target_min = current_user.fasting_target_min or 3.9
    
    for i, reading in enumerate(readings):
        is_anomaly = False
        if reading.context == "fasting" and reading.value > (current_user.fasting_target_max or 7.0):
            is_anomaly = True
        elif reading.context == "post_meal" and reading.value > target_max:
            is_anomaly = True
        elif reading.value < target_min:
            is_anomaly = True
        
        if is_anomaly:
            anomalies.append({
                "index": i,
                "date": reading.timestamp.isoformat(),
                "value": reading.value,
                "reason": "超出目标范围"
            })
    
    # 趋势方向判断
    if abs(slope) < 0.01:
        trend_direction = "stable"
        trend_text = "平稳"
    elif slope > 0:
        trend_direction = "up"
        trend_text = "上升"
    else:
        trend_direction = "down"
        trend_text = "下降"
    
    # 计算平均变化
    if len(values) >= 2:
        first_half = values[:len(values)//2]
        second_half = values[len(values)//2:]
        first_avg = sum(first_half) / len(first_half)
        second_avg = sum(second_half) / len(second_half)
        average_change = second_avg - first_avg
    else:
        average_change = 0
    
    # 生成解读
    interpretation = f"近{days}天血糖数据呈{trend_text}趋势"
    if abs(average_change) > 0.1:
        interpretation += f"，平均变化{abs(average_change):.2f} mmol/L"
    if anomalies:
        interpretation += f"，发现{len(anomalies)}个异常值"
    
    return {
        "trend": trend_direction,
        "trend_direction": trend_direction,
        "trend_text": trend_text,
        "average_change": round(average_change, 2),
        "interpretation": interpretation,
        "chart_data": {
            "dates": [d.isoformat() for d in dates],
            "values": values,
            "trend_line": trend_line,
            "anomalies": anomalies
        },
        "stats": {
            "average": round(sum(values) / len(values), 2),
            "max": max(values),
            "min": min(values),
            "std_dev": round(statistics.stdev(values) if len(values) > 1 else 0, 2)
        }
    }


# ==================== 提醒功能API ====================

@router.get("/users/reminders")
async def get_reminders(
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """获取用户的所有提醒"""
    reminders = db.query(models.Reminder).filter(
        models.Reminder.user_id == current_user.id
    ).order_by(models.Reminder.created_at.desc()).all()
    
    return {
        "reminders": [
            {
                "id": r.id,
                "reminder_type": r.reminder_type,
                "title": r.title,
                "content": r.content,
                "reminder_time": r.reminder_time,
                "reminder_date": r.reminder_date.isoformat() if r.reminder_date else None,
                "repeat_type": r.repeat_type,
                "repeat_days": r.repeat_days.split(",") if r.repeat_days else [],
                "enabled": r.enabled,
                "completed": r.completed,
                "completed_at": r.completed_at.isoformat() if r.completed_at else None,
                "created_at": r.created_at.isoformat()
            }
            for r in reminders
        ]
    }


@router.post("/users/reminders")
async def create_reminder(
    reminder_data: dict,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """创建提醒"""
    reminder = models.Reminder(
        user_id=current_user.id,
        reminder_type=reminder_data.get("reminder_type"),
        title=reminder_data.get("title"),
        content=reminder_data.get("content", ""),
        reminder_time=reminder_data.get("reminder_time"),
        reminder_date=datetime.fromisoformat(reminder_data["reminder_date"]) if reminder_data.get("reminder_date") else None,
        repeat_type=reminder_data.get("repeat_type", "daily"),
        repeat_days=",".join(reminder_data.get("repeat_days", [])) if reminder_data.get("repeat_days") else None,
        enabled=reminder_data.get("enabled", True)
    )
    
    db.add(reminder)
    db.commit()
    db.refresh(reminder)
    
    return {
        "id": reminder.id,
        "message": "提醒创建成功"
    }


@router.put("/users/reminders/{reminder_id}")
async def update_reminder(
    reminder_id: int,
    reminder_data: dict,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """更新提醒"""
    reminder = db.query(models.Reminder).filter(
        models.Reminder.id == reminder_id,
        models.Reminder.user_id == current_user.id
    ).first()
    
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")
    
    if "title" in reminder_data:
        reminder.title = reminder_data["title"]
    if "content" in reminder_data:
        reminder.content = reminder_data["content"]
    if "reminder_time" in reminder_data:
        reminder.reminder_time = reminder_data["reminder_time"]
    if "reminder_date" in reminder_data:
        reminder.reminder_date = datetime.fromisoformat(reminder_data["reminder_date"]) if reminder_data["reminder_date"] else None
    if "repeat_type" in reminder_data:
        reminder.repeat_type = reminder_data["repeat_type"]
    if "repeat_days" in reminder_data:
        reminder.repeat_days = ",".join(reminder_data["repeat_days"]) if reminder_data["repeat_days"] else None
    if "enabled" in reminder_data:
        reminder.enabled = reminder_data["enabled"]
    if "completed" in reminder_data:
        reminder.completed = reminder_data["completed"]
        if reminder_data["completed"]:
            reminder.completed_at = datetime.utcnow()
        else:
            reminder.completed_at = None
    
    db.commit()
    db.refresh(reminder)
    
    return {"message": "更新成功"}


@router.delete("/users/reminders/{reminder_id}")
async def delete_reminder(
    reminder_id: int,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """删除提醒"""
    reminder = db.query(models.Reminder).filter(
        models.Reminder.id == reminder_id,
        models.Reminder.user_id == current_user.id
    ).first()
    
    if not reminder:
        raise HTTPException(status_code=404, detail="提醒不存在")
    
    db.delete(reminder)
    db.commit()
    
    return {"message": "删除成功"}


# ==================== 数据导出API ====================

@router.get("/users/export/csv")
async def export_glucose_csv(
    days: int = 90,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """导出血糖数据为CSV"""
    from datetime import datetime, timedelta
    from fastapi.responses import Response
    import urllib.parse
    
    try:
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=days)
        
        readings = db.query(models.GlucoseReading).filter(
            models.GlucoseReading.user_id == current_user.id,
            models.GlucoseReading.timestamp >= start_date,
            models.GlucoseReading.timestamp <= end_date
        ).order_by(models.GlucoseReading.timestamp.desc()).all()
        
        # 创建CSV（使用StringIO先构建内容，然后转换为UTF-8-BOM编码）
        csv_string = io.StringIO()
        writer = csv.writer(csv_string)
        
        # 写入表头
        writer.writerow(["时间", "血糖值", "单位", "测量类型", "餐次", "风险等级", "备注"])
        
        # 写入数据
        for reading in readings:
            writer.writerow([
                reading.timestamp.strftime("%Y-%m-%d %H:%M:%S"),
                reading.value,
                reading.unit or "mmol/L",
                reading.context or "",
                reading.meal_type or "",
                reading.risk_level or "",
                reading.notes or ""
            ])
        
        # 获取CSV字符串内容
        csv_content_str = csv_string.getvalue()
        csv_string.close()
        
        # 转换为UTF-8-BOM编码（添加BOM以支持Excel正确显示中文）
        csv_content = '\ufeff' + csv_content_str
        csv_content = csv_content.encode('utf-8-sig')
        
        # 编码文件名以支持中文
        filename = f"DigiGlucose_血糖数据_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.csv"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        return Response(
            content=csv_content,
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "text/csv; charset=utf-8"
            }
        )
    except Exception as e:
        print(f"Error exporting CSV: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/users/export/excel")
async def export_glucose_excel(
    days: int = 90,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """导出血糖数据为Excel"""
    from datetime import datetime, timedelta
    from fastapi.responses import Response
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    import urllib.parse
    
    try:
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=days)
        
        readings = db.query(models.GlucoseReading).filter(
            models.GlucoseReading.user_id == current_user.id,
            models.GlucoseReading.timestamp >= start_date,
            models.GlucoseReading.timestamp <= end_date
        ).order_by(models.GlucoseReading.timestamp.desc()).all()
        
        # 创建Excel工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "血糖数据"
        
        # 设置表头
        headers = ["时间", "血糖值", "单位", "测量类型", "餐次", "风险等级", "备注"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 写入数据
        for row, reading in enumerate(readings, 2):
            ws.cell(row=row, column=1, value=reading.timestamp.strftime("%Y-%m-%d %H:%M:%S"))
            ws.cell(row=row, column=2, value=reading.value)
            ws.cell(row=row, column=3, value=reading.unit or "mmol/L")
            ws.cell(row=row, column=4, value=reading.context or "")
            ws.cell(row=row, column=5, value=reading.meal_type or "")
            ws.cell(row=row, column=6, value=reading.risk_level or "")
            ws.cell(row=row, column=7, value=reading.notes or "")
        
        # 调整列宽
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 30
        
        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        excel_content = output.getvalue()
        output.close()
        
        # 编码文件名以支持中文
        filename = f"DigiGlucose_血糖数据_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.xlsx"
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        return Response(
            content=excel_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )
    except Exception as e:
        print(f"Error exporting Excel: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")


@router.get("/users/export/pdf")
async def export_glucose_pdf(
    days: int = 90,
    current_user: models.User = Depends(auth.get_current_active_user),
    db: Session = Depends(database.get_db)
):
    """导出血糖数据为PDF报告"""
    from datetime import datetime, timedelta
    from fastapi.responses import Response
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    import platform
    import os
    
    try:
        # 注册中文字体
        # Windows系统字体路径
        if platform.system() == 'Windows':
            # 尝试注册常见的中文字体
            font_paths = [
                r'C:\Windows\Fonts\simsun.ttc',  # 宋体
                r'C:\Windows\Fonts\simhei.ttf',   # 黑体
                r'C:\Windows\Fonts\msyh.ttc',    # 微软雅黑
                r'C:\Windows\Fonts\simkai.ttf',   # 楷体
            ]
            
            chinese_font_name = 'SimSun'  # 默认使用宋体
            font_registered = False
            
            for font_path in font_paths:
                if os.path.exists(font_path):
                    try:
                        if font_path.endswith('.ttc'):
                            # TTC文件需要指定字体索引
                            pdfmetrics.registerFont(TTFont('SimSun', font_path, subfontIndex=0))
                        else:
                            pdfmetrics.registerFont(TTFont('SimSun', font_path))
                        font_registered = True
                        print(f"成功注册中文字体: {font_path}")
                        break
                    except Exception as e:
                        print(f"注册字体失败 {font_path}: {str(e)}")
                        continue
            
            if not font_registered:
                # 如果无法注册系统字体，使用reportlab的CJK支持（需要安装reportlab的CJK字体包）
                print("警告: 无法注册系统中文字体，尝试使用备用方案")
                chinese_font_name = 'Helvetica'  # 备用方案，但可能无法显示中文
        else:
            # Linux/Mac系统，尝试使用系统字体或CJK字体
            chinese_font_name = 'Helvetica'
            print("非Windows系统，使用默认字体（可能不支持中文）")
        
        end_date = datetime.utcnow()
        start_date = end_date - timedelta(days=days)
        
        readings = db.query(models.GlucoseReading).filter(
            models.GlucoseReading.user_id == current_user.id,
            models.GlucoseReading.timestamp >= start_date,
            models.GlucoseReading.timestamp <= end_date
        ).order_by(models.GlucoseReading.timestamp.desc()).limit(1000).all()  # 限制1000条避免过大
        
        # 创建PDF
        output = io.BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # 创建支持中文的样式
        chinese_normal_style = ParagraphStyle(
            'ChineseNormal',
            parent=styles['Normal'],
            fontName=chinese_font_name,
            fontSize=10,
            leading=14
        )
        
        chinese_title_style = ParagraphStyle(
            'ChineseTitle',
            parent=styles['Heading1'],
            fontName=chinese_font_name,
            fontSize=24,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=30
        )
        
        chinese_heading_style = ParagraphStyle(
            'ChineseHeading',
            parent=styles['Heading2'],
            fontName=chinese_font_name,
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12
        )
        
        # 标题
        story.append(Paragraph("血糖数据报告", chinese_title_style))
        
        # 报告信息
        info_text = f"<b>用户：</b>{current_user.username}<br/>"
        info_text += f"<b>时间范围：</b>{start_date.strftime('%Y-%m-%d')} 至 {end_date.strftime('%Y-%m-%d')}<br/>"
        info_text += f"<b>记录数量：</b>{len(readings)}条"
        story.append(Paragraph(info_text, chinese_normal_style))
        story.append(Spacer(1, 0.3*inch))
        
        # 统计数据
        if readings:
            values = [r.value for r in readings]
            stats_text = f"<b>平均值：</b>{sum(values)/len(values):.2f} mmol/L<br/>"
            stats_text += f"<b>最高值：</b>{max(values):.2f} mmol/L<br/>"
            stats_text += f"<b>最低值：</b>{min(values):.2f} mmol/L"
            story.append(Paragraph(stats_text, chinese_normal_style))
            story.append(Spacer(1, 0.3*inch))
        
        # 数据表格
        if readings:
            data = [["时间", "血糖值", "单位", "测量类型", "风险等级"]]
            for reading in readings[:100]:  # 限制显示前100条
                # 转换测量类型为中文
                context_map = {
                    'fasting': '空腹',
                    'post_meal': '餐后',
                    'pre_meal': '餐前',
                    'random': '随机'
                }
                context_cn = context_map.get(reading.context, reading.context or "-")
                
                # 转换风险等级为中文
                risk_map = {
                    'normal': '正常',
                    'high': '偏高',
                    'low': '偏低',
                    'critical_high': '严重偏高',
                    'critical_low': '严重偏低'
                }
                risk_cn = risk_map.get(reading.risk_level, reading.risk_level or "-")
                
                data.append([
                    reading.timestamp.strftime("%Y-%m-%d %H:%M"),
                    f"{reading.value:.1f}",
                    reading.unit or "mmol/L",
                    context_cn,
                    risk_cn
                ])
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), chinese_font_name),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), chinese_font_name),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
            ]))
            story.append(table)
        
        # 生成PDF
        doc.build(story)
        output.seek(0)
        pdf_content = output.getvalue()
        output.close()
        
        # 编码文件名以支持中文
        filename = f"DigiGlucose_血糖数据_{start_date.strftime('%Y%m%d')}-{end_date.strftime('%Y%m%d')}.pdf"
        import urllib.parse
        encoded_filename = urllib.parse.quote(filename.encode('utf-8'))
        
        return Response(
            content=pdf_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
                "Content-Type": "application/pdf"
            }
        )
    except Exception as e:
        print(f"Error exporting PDF: {str(e)}")
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"导出失败: {str(e)}")

